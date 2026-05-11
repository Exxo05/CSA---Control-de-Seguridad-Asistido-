# gui/screens/exportar.py — Exportar/imprimir informes
import tkinter as tk
from tkinter import messagebox, filedialog
import os, threading
from datetime import datetime

from servicios.db   import get_connection, hacer_backup
from servicios.sesion import nombre as sesion_nombre
from gui.styles import (
    POLICE_BLUE, WHITE, GRAY_BG, GRAY_BORDER, DARK_TEXT, GRAY_TEXT,
    BAJA_BG, BAJA_FG, FONT_NORMAL, FONT_SUBTITLE, FONT_SMALL,
    make_button, make_header, PAD_X, PAD_Y
)


class ExportarScreen(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg=GRAY_BG)
        self._build()

    def _build(self):
        make_header(self, "📄  Exportar e Informes")

        body = tk.Frame(self, bg=GRAY_BG)
        body.pack(fill="both", expand=True, padx=PAD_X, pady=16)

        # ── Informe de turno ──────────────────────────────────────
        self._seccion(body, "📋  Informe de turno",
                      "Genera un PDF con todos los incidentes del turno actual.",
                      [("🖨️  Generar informe de turno PDF",  self._informe_turno, "primary"),
                       ("📊  Exportar incidentes a Excel",   self._exportar_excel, "success")])

        # ── Backup ────────────────────────────────────────────────
        self._seccion(body, "💾  Copia de seguridad",
                      "Hace una copia de la base de datos en datos/backups/",
                      [("💾  Hacer backup ahora",           self._hacer_backup, "info")])

        # ── Parte individual ──────────────────────────────────────
        self._seccion(body, "📝  Parte de incidente individual",
                      "Introduce el ID del incidente para generar su parte en PDF.",
                      [], extra=self._build_parte_individual)

    def _seccion(self, parent, titulo, desc, botones, extra=None):
        card = tk.Frame(parent, bg=WHITE,
                        highlightbackground=GRAY_BORDER, highlightthickness=1)
        card.pack(fill="x", pady=(0,12))
        hdr = tk.Frame(card, bg="#F8FAFC")
        hdr.pack(fill="x")
        tk.Label(hdr, text=titulo, font=FONT_SUBTITLE,
                 bg="#F8FAFC", fg=DARK_TEXT, pady=8, padx=PAD_X).pack(anchor="w")
        tk.Label(hdr, text=desc, font=FONT_SMALL,
                 bg="#F8FAFC", fg=GRAY_TEXT, padx=PAD_X).pack(anchor="w", pady=(0,6))
        tk.Frame(card, bg=GRAY_BORDER, height=1).pack(fill="x")
        inner = tk.Frame(card, bg=WHITE)
        inner.pack(fill="x", padx=PAD_X, pady=PAD_Y)
        for txt, cmd, style in botones:
            make_button(inner, txt, cmd, style).pack(side="left", padx=(0,8))
        if extra:
            extra(inner)

    def _build_parte_individual(self, parent):
        row = tk.Frame(parent, bg=WHITE)
        row.pack(fill="x", pady=(12,0))
        tk.Label(row, text="ID del incidente:", font=FONT_SMALL,
                 bg=WHITE, fg=GRAY_TEXT).pack(side="left")
        self.ent_id = tk.Entry(row, font=FONT_NORMAL, relief="solid",
                                bd=1, bg=WHITE, width=8)
        self.ent_id.pack(side="left", padx=(4,8), ipady=4)
        make_button(row, "📄  Generar parte PDF",
                    self._parte_individual, "warning").pack(side="left")

    # ── Acciones ──────────────────────────────────────────────────
    def _hacer_backup(self):
        ruta = hacer_backup()
        if ruta:
            messagebox.showinfo("✅ Backup completado",
                                f"Copia guardada en:\n{ruta}")
        else:
            messagebox.showerror("Error","No se pudo completar el backup.")

    def _exportar_excel(self):
        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel","*.xlsx")],
            initialfile=f"CSA_incidentes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        )
        if not ruta: return
        threading.Thread(target=self._worker_excel, args=(ruta,), daemon=True).start()

    def _worker_excel(self, ruta):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Incidentes"

            # Cabecera
            headers = ["ID","Tipo","Descripción","Fecha","Zona","Dirección",
                       "Estado","Lat","Lon"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0A1F44")
                cell.alignment = Alignment(horizontal="center")

            conn = get_connection()
            c    = conn.cursor()
            c.execute("""SELECT id,tipo,descripcion,fecha,zona,direccion,
                                estado,lat,lon FROM incidentes ORDER BY id DESC""")
            for row_i, row in enumerate(c.fetchall(), 2):
                for col_i, val in enumerate(row, 1):
                    ws.cell(row=row_i, column=col_i, value=val)
            conn.close()

            # Autoajustar columnas
            for col in ws.columns:
                max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len+2, 50)

            wb.save(ruta)
            self.after(0, lambda: messagebox.showinfo(
                "✅ Exportado", f"Archivo guardado:\n{ruta}"))
        except ImportError:
            self.after(0, lambda: messagebox.showerror(
                "Librería faltante",
                "Instala openpyxl:\n  pip install openpyxl"))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", str(e)))

    def _informe_turno(self):
        ruta = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF","*.pdf")],
            initialfile=f"Informe_turno_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
        )
        if not ruta: return
        threading.Thread(target=self._worker_pdf_turno,
                         args=(ruta,), daemon=True).start()

    def _worker_pdf_turno(self, ruta):
        try:
            self._generar_pdf_turno(ruta)
            self.after(0, lambda: messagebox.showinfo(
                "✅ PDF generado", f"Informe guardado en:\n{ruta}"))
        except ImportError:
            self.after(0, lambda: messagebox.showerror(
                "Librería faltante",
                "Instala reportlab:\n  pip install reportlab"))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error PDF", str(e)))

    def _generar_pdf_turno(self, ruta):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                         Paragraph, Spacer)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units  import cm

        doc    = SimpleDocTemplate(ruta, pagesize=A4,
                                   topMargin=2*cm, bottomMargin=2*cm,
                                   leftMargin=2*cm, rightMargin=2*cm)
        styles = getSampleStyleSheet()
        story  = []

        # Título
        titulo_style = ParagraphStyle("titulo", fontSize=16, fontName="Helvetica-Bold",
                                       textColor=colors.HexColor("#0A1F44"),
                                       spaceAfter=6)
        story.append(Paragraph("INFORME DE TURNO — CSA", titulo_style))
        story.append(Paragraph(
            f"Policía Local de Alcalá de Henares  |  "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  "
            f"Operador: {sesion_nombre()}",
            styles["Normal"]))
        story.append(Spacer(1, 0.5*cm))

        # Datos
        conn = get_connection()
        c    = conn.cursor()
        hoy  = datetime.now().strftime("%Y-%m-%d")
        c.execute("""SELECT id,tipo,zona,fecha,estado FROM incidentes
                     WHERE fecha LIKE ? ORDER BY fecha""", (f"{hoy}%",))
        rows = c.fetchall()
        conn.close()

        story.append(Paragraph(f"Incidentes del día ({len(rows)} total):",
                               styles["Heading2"]))
        story.append(Spacer(1, 0.3*cm))

        tabla_data = [["ID","Tipo","Zona","Hora","Estado"]]
        for r in rows:
            tabla_data.append([
                str(r[0]),
                (r[1][:40]+"…" if r[1] and len(r[1])>40 else r[1] or "—"),
                r[2] or "—",
                r[3][11:16] if r[3] else "—",
                r[4] or "—"
            ])

        t = Table(tabla_data, colWidths=[1.2*cm,7*cm,4*cm,2*cm,2.5*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0), colors.HexColor("#0A1F44")),
            ("TEXTCOLOR",    (0,0), (-1,0), colors.white),
            ("FONTNAME",     (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),
             [colors.HexColor("#F8FAFC"), colors.white]),
            ("GRID",         (0,0), (-1,-1), 0.3, colors.HexColor("#DDE3ED")),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
            ("PADDING",      (0,0), (-1,-1), 4),
        ]))
        story.append(t)
        doc.build(story)

    def _parte_individual(self):
        inc_id = self.ent_id.get().strip()
        if not inc_id.isdigit():
            messagebox.showwarning("ID inválido","Introduce un número de ID válido.")
            return
        ruta = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF","*.pdf")],
            initialfile=f"Parte_incidente_{inc_id}.pdf"
        )
        if not ruta: return
        threading.Thread(target=self._worker_parte,
                         args=(int(inc_id), ruta), daemon=True).start()

    def _worker_parte(self, inc_id, ruta):
        try:
            conn = get_connection()
            c    = conn.cursor()
            c.execute("SELECT * FROM incidentes WHERE id=?", (inc_id,))
            inc  = c.fetchone()
            c.execute("""SELECT texto, fecha FROM notas_incidente
                         WHERE incidente_id=? ORDER BY fecha""", (inc_id,))
            notas = c.fetchall()
            conn.close()

            if not inc:
                self.after(0, lambda: messagebox.showerror(
                    "No encontrado",f"No existe el incidente #{inc_id}"))
                return

            self._generar_parte_pdf(ruta, inc, notas)
            self.after(0, lambda: messagebox.showinfo(
                "✅ Parte generado", f"Guardado en:\n{ruta}"))
        except ImportError:
            self.after(0, lambda: messagebox.showerror(
                "Librería faltante","pip install reportlab"))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Error", str(e)))

    def _generar_parte_pdf(self, ruta, inc, notas):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph,
                                         Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units  import cm

        doc    = SimpleDocTemplate(ruta, pagesize=A4,
                                   topMargin=2*cm, bottomMargin=2*cm,
                                   leftMargin=2.5*cm, rightMargin=2.5*cm)
        styles = getSampleStyleSheet()
        story  = []

        azul   = colors.HexColor("#0A1F44")
        titulo = ParagraphStyle("t", fontSize=18, fontName="Helvetica-Bold",
                                textColor=azul, spaceAfter=4)
        campo  = ParagraphStyle("c", fontSize=10, fontName="Helvetica-Bold",
                                textColor=azul, spaceBefore=8)
        valor  = ParagraphStyle("v", fontSize=10, fontName="Helvetica",
                                leftIndent=12)

        story.append(Paragraph("PARTE DE INCIDENTE POLICIAL", titulo))
        story.append(Paragraph(
            f"Policía Local de Alcalá de Henares  —  CSA v2.0",
            styles["Normal"]))
        story.append(HRFlowable(width="100%", thickness=2,
                                color=azul, spaceAfter=12))

        campos = [
            ("Número de expediente:", f"#{inc[0]}"),
            ("Tipo / Categoría:",     inc[1] or "—"),
            ("Fecha y hora:",         inc[3] or "—"),
            ("Zona / Barrio:",        inc[4] or "—"),
            ("Dirección exacta:",     inc[5] or "—"),
            ("Estado actual:",        inc[6] or "—"),
            ("Descripción:",          inc[2] or "—"),
            ("Operador/turno:",       inc[10] or "—" if len(inc)>10 else "—"),
        ]
        for lbl, val in campos:
            story.append(Paragraph(lbl, campo))
            story.append(Paragraph(str(val), valor))

        if notas:
            story.append(Spacer(1, 0.5*cm))
            story.append(HRFlowable(width="100%", thickness=1,
                                    color=colors.lightgrey))
            story.append(Paragraph("NOTAS / ACTUALIZACIONES:", campo))
            for nota in notas:
                story.append(Paragraph(
                    f"[{nota[1][:19]}]  {nota[0]}", valor))

        story.append(Spacer(1, 2*cm))
        story.append(Paragraph("Firma del operador: ____________________________", valor))
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f"Documento generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  {sesion_nombre()}",
            ParagraphStyle("pie", fontSize=8, textColor=colors.grey)))

        doc.build(story)
